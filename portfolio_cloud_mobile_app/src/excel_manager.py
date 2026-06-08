from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.portfolio_calculator import (
    CAPITAL_FLOW_COLUMNS,
    SNAPSHOT_COLUMNS,
    TRANSACTION_COLUMNS,
    empty_capital_flows,
    empty_prices,
    empty_snapshots,
    empty_transactions,
    normalize_capital_flows,
    normalize_holdings,
    normalize_snapshots,
    sample_holdings,
    infer_major_asset_class,
    normalize_sub_asset_class,
)
from src.style_config import ASSET_CLASS_COLORS, MAJOR_ASSET_CLASS_COLORS


DISCLOSURE_COLUMNS = [
    "저장일시",
    "시장",
    "티커_또는_종목코드",
    "종목명",
    "공시일",
    "공시유형",
    "공시제목",
    "공시원문URL",
    "공시ID",
    "요약",
    "중요도",
    "처리상태",
]

DISCLOSURE_LOG_COLUMNS = [
    "실행일시",
    "조회모드",
    "조회시작일",
    "조회종료일",
    "조회대상종목수",
    "조회대상종목목록",
    "성공종목수",
    "실패종목수",
    "조회된공시수",
    "신규저장공시수",
    "중복제외공시수",
    "필터후표시공시수",
    "오류메시지",
    "상세로그",
]

DISCLOSURE_WATCHLIST_COLUMNS = [
    "시장",
    "티커_또는_종목코드",
    "종목명",
    "추적상태",
    "추가방식",
    "추가일시",
    "메모",
]

SHEET_NAMES = [
    "holdings",
    "prices",
    "calculated",
    "dashboard",
    "settings",
    "transactions",
    "capital_flows",
    "portfolio_snapshots",
    "disclosures",
    "disclosure_logs",
    "disclosure_watchlist",
]


def ensure_workbook_exists(workbook_path: Path) -> None:
    if workbook_path.exists():
        _ensure_required_sheets(workbook_path)
        return
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    save_full_workbook(
        workbook_path=workbook_path,
        holdings=sample_holdings(),
        prices=empty_prices(),
        calculated=pd.DataFrame(),
        settings=default_settings(),
        create_backup=False,
    )


def _ensure_required_sheets(workbook_path: Path) -> None:
    try:
        workbook = load_workbook(workbook_path)
    except Exception:
        return
    changed = False
    required_headers = {
        "holdings": ["상위자산군", "세부자산군", "row_id"],
        "transactions": ["상위자산군", "세부자산군"],
        "capital_flows": CAPITAL_FLOW_COLUMNS,
        "portfolio_snapshots": SNAPSHOT_COLUMNS,
        "settings": ["설정", "값"],
        "disclosures": DISCLOSURE_COLUMNS,
        "disclosure_logs": DISCLOSURE_LOG_COLUMNS,
        "disclosure_watchlist": DISCLOSURE_WATCHLIST_COLUMNS,
    }
    for sheet_name, headers in required_headers.items():
        if sheet_name not in workbook.sheetnames:
            ws = workbook.create_sheet(sheet_name)
            for col_index, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_index, value=header)
            changed = True
            continue
        ws = workbook[sheet_name]
        existing_headers = [cell.value for cell in ws[1]]
        for header in headers:
            if header not in existing_headers:
                ws.cell(row=1, column=ws.max_column + 1, value=header)
                existing_headers.append(header)
                changed = True
    if changed:
        for sheet_name in workbook.sheetnames:
            _style_sheet(workbook[sheet_name])
        try:
            workbook.save(workbook_path)
        except PermissionError:
            pass


def default_settings() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["기본통화", "KRW"],
            ["가격표시문구", "최신 조회 가능 가격 기준"],
            ["주식 벤치마크 티커", "VT"],
            ["채권 벤치마크 티커", "BND"],
            ["금 벤치마크 티커", "GLD"],
            ["주식 비중", "60"],
            ["채권 비중", "30"],
            ["금 비중", "10"],
            ["OpenDART API Key", ""],
            ["OpenAI API Key", ""],
            ["요약 모델명", "gpt-4o-mini"],
            ["SEC User-Agent", "Personal Portfolio Disclosure Tracker contact@example.com"],
            ["dart_api_key", ""],
            ["openai_api_key", ""],
            ["summary_model", "gpt-4o-mini"],
            ["sec_user_agent", "Personal Portfolio Disclosure Tracker contact@example.com"],
            ["last_disclosure_refresh_datetime", ""],
            ["disclosure_first_refresh_completed", "False"],
            ["openai_available", "False"],
            ["openai_error_type", ""],
            ["종목당 최대 공시 조회 건수", "30"],
            ["주식 색상", MAJOR_ASSET_CLASS_COLORS["주식"]],
            ["ETF 색상", ASSET_CLASS_COLORS["ETF"]],
            ["개별주 색상", ASSET_CLASS_COLORS["개별주"]],
            ["미국채권 색상", ASSET_CLASS_COLORS["미국채권"]],
            ["국내채권 색상", ASSET_CLASS_COLORS["국내채권"]],
            ["한국리츠 색상", ASSET_CLASS_COLORS["한국리츠"]],
            ["암호화폐 색상", ASSET_CLASS_COLORS["암호화폐"]],
            ["달러 색상", ASSET_CLASS_COLORS["달러"]],
        ],
        columns=["설정", "값"],
    )


def normalize_settings(settings: pd.DataFrame | None) -> pd.DataFrame:
    defaults = default_settings()
    if settings is None or settings.empty or "설정" not in settings.columns:
        return defaults
    normalized = settings.copy()
    if "값" not in normalized.columns:
        normalized["값"] = ""
    normalized["설정"] = normalized["설정"].fillna("").astype(str)
    normalized["값"] = normalized["값"].fillna("").astype(str)
    existing_names = set(normalized["설정"])
    if "금 벤치마크 티커" not in existing_names and "채권 비중" in existing_names:
        old_bond_weight = normalized.loc[normalized["설정"] == "채권 비중", "값"].astype(str).str.strip()
        if not old_bond_weight.empty and old_bond_weight.iloc[0] == "40":
            normalized.loc[normalized["설정"] == "채권 비중", "값"] = "30"
    merged = pd.concat(
        [
            normalized[normalized["설정"].astype(str).str.strip() != ""][["설정", "값"]],
            defaults[~defaults["설정"].isin(set(normalized["설정"]))],
        ],
        ignore_index=True,
    )
    return merged.drop_duplicates(subset=["설정"], keep="first").reset_index(drop=True)


def settings_to_dict(settings: pd.DataFrame | None) -> dict[str, str]:
    normalized = normalize_settings(settings)
    values = {str(row["설정"]): str(row["값"]) for _, row in normalized.iterrows()}
    alias_pairs = [
        ("OpenDART API Key", "dart_api_key"),
        ("OpenAI API Key", "openai_api_key"),
        ("요약 모델명", "summary_model"),
        ("SEC User-Agent", "sec_user_agent"),
    ]
    for display_key, internal_key in alias_pairs:
        if values.get(display_key) and not values.get(internal_key):
            values[internal_key] = values[display_key]
        if values.get(internal_key) and not values.get(display_key):
            values[display_key] = values[internal_key]
    return values


def load_holdings(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    _backup_before_schema_migration(workbook_path)
    try:
        return normalize_holdings(pd.read_excel(workbook_path, sheet_name="holdings"))
    except Exception:
        return sample_holdings()


def _backup_before_schema_migration(workbook_path: Path) -> None:
    if not workbook_path.exists():
        return
    workbook = None
    try:
        workbook = load_workbook(workbook_path, read_only=True)
        if "holdings" not in workbook.sheetnames:
            return
        ws = workbook["holdings"]
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        header_set = set(headers)
        required = {"새빛_보유수량", "희주_보유수량", "합산_보유수량", "row_id"}
        needs_backup = not required.issubset(header_set)
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = dict(zip(headers, row))
            if values.get("자산군") == "비트코인" or values.get("세부자산군") == "비트코인":
                needs_backup = True
                break
        if not needs_backup:
            return
        backup_dir = workbook_path.parent / "data" / "backup"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"portfolio_before_schema_migration_{timestamp}.xlsx"
        shutil.copy2(workbook_path, backup_path)
    except Exception:
        return
    finally:
        if workbook is not None:
            workbook.close()


def load_prices(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        return pd.read_excel(workbook_path, sheet_name="prices")
    except Exception:
        return empty_prices()


def load_settings(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        return normalize_settings(pd.read_excel(workbook_path, sheet_name="settings", dtype=str))
    except Exception:
        return default_settings()


def load_disclosures(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        disclosures = pd.read_excel(workbook_path, sheet_name="disclosures", dtype=str)
    except Exception:
        return pd.DataFrame(columns=DISCLOSURE_COLUMNS)
    for column in DISCLOSURE_COLUMNS:
        if column not in disclosures.columns:
            disclosures[column] = ""
    return disclosures[DISCLOSURE_COLUMNS].fillna("")


def load_disclosure_logs(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        logs = pd.read_excel(workbook_path, sheet_name="disclosure_logs", dtype=str)
    except Exception:
        return pd.DataFrame(columns=DISCLOSURE_LOG_COLUMNS)
    for column in DISCLOSURE_LOG_COLUMNS:
        if column not in logs.columns:
            logs[column] = ""
    return logs[DISCLOSURE_LOG_COLUMNS].fillna("")


def load_disclosure_watchlist(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        watchlist = pd.read_excel(workbook_path, sheet_name="disclosure_watchlist", dtype=str)
    except Exception:
        return pd.DataFrame(columns=DISCLOSURE_WATCHLIST_COLUMNS)
    for column in DISCLOSURE_WATCHLIST_COLUMNS:
        if column not in watchlist.columns:
            watchlist[column] = ""
    return watchlist[DISCLOSURE_WATCHLIST_COLUMNS].fillna("")


def load_calculated(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        calculated = pd.read_excel(workbook_path, sheet_name="calculated")
        if "원화 환산 평가금액" not in calculated.columns:
            return pd.DataFrame()
        return calculated
    except Exception:
        return pd.DataFrame()


def load_transactions(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        transactions = pd.read_excel(workbook_path, sheet_name="transactions", dtype={"티커 또는 종목코드": str})
        for column in TRANSACTION_COLUMNS:
            if column not in transactions.columns:
                transactions[column] = ""
        return transactions[TRANSACTION_COLUMNS]
    except Exception:
        return empty_transactions()


def load_capital_flows(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        return normalize_capital_flows(pd.read_excel(workbook_path, sheet_name="capital_flows"))
    except Exception:
        return empty_capital_flows()


def load_snapshots(workbook_path: Path) -> pd.DataFrame:
    ensure_workbook_exists(workbook_path)
    try:
        return normalize_snapshots(pd.read_excel(workbook_path, sheet_name="portfolio_snapshots"))
    except Exception:
        return empty_snapshots()


def create_backup(workbook_path: Path, backup_dir: Path) -> Path | None:
    if not workbook_path.exists():
        return None
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"portfolio_{timestamp}.xlsx"
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def save_holdings(workbook_path: Path, backup_dir: Path, holdings: pd.DataFrame) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    holdings = normalize_holdings(holdings)
    if "sort_order" in holdings.columns:
        holdings = holdings.sort_values("sort_order", kind="stable").reset_index(drop=True)
    prices = load_prices(workbook_path)
    calculated = load_calculated(workbook_path)
    transactions = load_transactions(workbook_path)
    capital_flows = load_capital_flows(workbook_path)
    snapshots = load_snapshots(workbook_path)
    disclosures = load_disclosures(workbook_path)
    save_full_workbook(workbook_path, holdings, prices, calculated, load_settings(workbook_path), transactions, capital_flows, snapshots, disclosures, create_backup=False)


def save_prices_and_calculated(
    workbook_path: Path,
    backup_dir: Path,
    holdings: pd.DataFrame,
    prices: pd.DataFrame,
    calculated: pd.DataFrame,
) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    transactions = load_transactions(workbook_path)
    capital_flows = load_capital_flows(workbook_path)
    snapshots = load_snapshots(workbook_path)
    disclosures = load_disclosures(workbook_path)
    save_full_workbook(workbook_path, holdings, prices, calculated, load_settings(workbook_path), transactions, capital_flows, snapshots, disclosures, create_backup=False)


def save_holdings_with_transaction(
    workbook_path: Path,
    backup_dir: Path,
    holdings: pd.DataFrame,
    transaction: pd.DataFrame,
    prices: pd.DataFrame,
    calculated: pd.DataFrame,
) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    existing_transactions = load_transactions(workbook_path)
    transactions = transaction.copy() if existing_transactions.empty else pd.concat([existing_transactions, transaction], ignore_index=True)
    capital_flows = load_capital_flows(workbook_path)
    snapshots = load_snapshots(workbook_path)
    disclosures = load_disclosures(workbook_path)
    save_full_workbook(workbook_path, holdings, prices, calculated, load_settings(workbook_path), transactions, capital_flows, snapshots, disclosures, create_backup=False)


def save_capital_flows(workbook_path: Path, backup_dir: Path, capital_flows: pd.DataFrame) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    save_full_workbook(
        workbook_path,
        load_holdings(workbook_path),
        load_prices(workbook_path),
        load_calculated(workbook_path),
        load_settings(workbook_path),
        load_transactions(workbook_path),
        capital_flows,
        load_snapshots(workbook_path),
        load_disclosures(workbook_path),
        create_backup=False,
    )


def save_snapshots(workbook_path: Path, backup_dir: Path, snapshots: pd.DataFrame) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    save_full_workbook(
        workbook_path,
        load_holdings(workbook_path),
        load_prices(workbook_path),
        load_calculated(workbook_path),
        load_settings(workbook_path),
        load_transactions(workbook_path),
        load_capital_flows(workbook_path),
        snapshots,
        load_disclosures(workbook_path),
        create_backup=False,
    )


def save_holdings_transaction_and_capital(
    workbook_path: Path,
    backup_dir: Path,
    holdings: pd.DataFrame,
    transaction: pd.DataFrame,
    prices: pd.DataFrame,
    calculated: pd.DataFrame,
    capital_flows: pd.DataFrame,
) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    existing_transactions = load_transactions(workbook_path)
    transactions = transaction.copy() if existing_transactions.empty else pd.concat([existing_transactions, transaction], ignore_index=True)
    snapshots = load_snapshots(workbook_path)
    disclosures = load_disclosures(workbook_path)
    save_full_workbook(workbook_path, holdings, prices, calculated, load_settings(workbook_path), transactions, capital_flows, snapshots, disclosures, create_backup=False)


def save_settings(workbook_path: Path, backup_dir: Path, settings: pd.DataFrame) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    save_full_workbook(
        workbook_path,
        load_holdings(workbook_path),
        load_prices(workbook_path),
        load_calculated(workbook_path),
        normalize_settings(settings),
        load_transactions(workbook_path),
        load_capital_flows(workbook_path),
        load_snapshots(workbook_path),
        load_disclosures(workbook_path),
        create_backup=False,
    )


def save_disclosures(
    workbook_path: Path,
    backup_dir: Path,
    disclosures: pd.DataFrame,
    settings: pd.DataFrame | None = None,
    disclosure_logs: pd.DataFrame | None = None,
    disclosure_watchlist: pd.DataFrame | None = None,
) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)
    save_full_workbook(
        workbook_path,
        load_holdings(workbook_path),
        load_prices(workbook_path),
        load_calculated(workbook_path),
        normalize_settings(settings if settings is not None else load_settings(workbook_path)),
        load_transactions(workbook_path),
        load_capital_flows(workbook_path),
        load_snapshots(workbook_path),
        normalize_disclosures(disclosures),
        normalize_disclosure_logs(disclosure_logs if disclosure_logs is not None else load_disclosure_logs(workbook_path)),
        normalize_disclosure_watchlist(disclosure_watchlist if disclosure_watchlist is not None else load_disclosure_watchlist(workbook_path)),
        create_backup=False,
    )


def save_disclosure_watchlist(workbook_path: Path, backup_dir: Path, watchlist: pd.DataFrame) -> None:
    save_disclosures(
        workbook_path,
        backup_dir,
        load_disclosures(workbook_path),
        load_settings(workbook_path),
        load_disclosure_logs(workbook_path),
        watchlist,
    )


def save_dashboard(
    workbook_path: Path,
    backup_dir: Path,
    summary: dict[str, float],
    asset_summary: pd.DataFrame,
    chart_paths: dict[str, Path],
) -> None:
    ensure_workbook_exists(workbook_path)
    create_backup(workbook_path, backup_dir)

    workbook = load_workbook(workbook_path)
    if "dashboard" in workbook.sheetnames:
        del workbook["dashboard"]
    ws = workbook.create_sheet("dashboard")

    ws["A1"] = "개인 자산관리 대시보드"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "요약"
    ws["A3"].font = Font(size=13, bold=True)

    for row_index, (key, value) in enumerate(summary.items(), start=4):
        ws.cell(row=row_index, column=1, value=key)
        ws.cell(row=row_index, column=2, value=float(value))

    start_row = 10
    ws.cell(row=start_row, column=1, value="세부자산군별 평가금액")
    ws.cell(row=start_row, column=1).font = Font(size=13, bold=True)
    for col_index, column in enumerate(asset_summary.columns, start=1):
        ws.cell(row=start_row + 1, column=col_index, value=column)
    for r, row in enumerate(asset_summary.itertuples(index=False), start=start_row + 2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    positions = {
        "asset_donut": "E3",
        "major_asset_donut": "E35",
        "etf_weight_bar": "E67",
        "individual_stock_weight_bar": "E99",
        "holding_value_bar": "E131",
        "asset_value_bar": "E163",
        "account_value_bar": "E195",
        "saebit_asset_donut": "E227",
        "heeju_asset_donut": "E259",
        "saebit_major_asset_donut": "E291",
        "heeju_major_asset_donut": "E323",
    }
    for name, path in chart_paths.items():
        if path.exists():
            image = ExcelImage(str(path))
            image.width = 760
            image.height = 450
            ws.add_image(image, positions.get(name, "E3"))

    _style_sheet(ws)
    workbook.save(workbook_path)


def save_full_workbook(
    workbook_path: Path,
    holdings: pd.DataFrame,
    prices: pd.DataFrame,
    calculated: pd.DataFrame,
    settings: pd.DataFrame,
    transactions: pd.DataFrame | None = None,
    capital_flows: pd.DataFrame | None = None,
    snapshots: pd.DataFrame | None = None,
    disclosures: pd.DataFrame | None = None,
    disclosure_logs: pd.DataFrame | None = None,
    disclosure_watchlist: pd.DataFrame | None = None,
    create_backup: bool = False,
) -> None:
    if create_backup:
        backup_dir = workbook_path.parent / "data" / "backup"
        globals()["create_backup"](workbook_path, backup_dir)
    if disclosure_logs is None and workbook_path.exists():
        disclosure_logs = load_disclosure_logs(workbook_path)
    if disclosure_watchlist is None and workbook_path.exists():
        disclosure_watchlist = load_disclosure_watchlist(workbook_path)

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        normalized_holdings = normalize_holdings(holdings)
        if "sort_order" in normalized_holdings.columns:
            normalized_holdings = normalized_holdings.sort_values("sort_order", kind="stable").reset_index(drop=True)
        normalized_holdings.to_excel(writer, sheet_name="holdings", index=False)
        prices.to_excel(writer, sheet_name="prices", index=False)
        calculated.to_excel(writer, sheet_name="calculated", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="dashboard", index=False)
        normalize_settings(settings).to_excel(writer, sheet_name="settings", index=False)
        _normalize_transactions(transactions).to_excel(writer, sheet_name="transactions", index=False)
        normalize_capital_flows(capital_flows).to_excel(writer, sheet_name="capital_flows", index=False)
        normalize_snapshots(snapshots).to_excel(writer, sheet_name="portfolio_snapshots", index=False)
        normalize_disclosures(disclosures).to_excel(writer, sheet_name="disclosures", index=False)
        normalize_disclosure_logs(disclosure_logs).to_excel(writer, sheet_name="disclosure_logs", index=False)
        normalize_disclosure_watchlist(disclosure_watchlist).to_excel(writer, sheet_name="disclosure_watchlist", index=False)

    workbook = load_workbook(workbook_path)
    for sheet_name in workbook.sheetnames:
        _style_sheet(workbook[sheet_name])
    _format_symbol_columns_as_text(workbook)
    workbook.save(workbook_path)


def _normalize_transactions(transactions: pd.DataFrame | None) -> pd.DataFrame:
    if transactions is None or transactions.empty:
        return empty_transactions()
    normalized = transactions.copy()
    if "자산군" in normalized.columns:
        normalized["세부자산군"] = normalized["자산군"]
    elif "세부자산군" not in normalized.columns:
        normalized["세부자산군"] = ""
    normalized["세부자산군"] = normalized["세부자산군"].map(normalize_sub_asset_class)
    normalized["자산군"] = normalized["세부자산군"]
    normalized["상위자산군"] = normalized["세부자산군"].map(infer_major_asset_class)
    for column in TRANSACTION_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = ""
    normalized["티커 또는 종목코드"] = normalized["티커 또는 종목코드"].fillna("").astype(str)
    return normalized[TRANSACTION_COLUMNS]


def normalize_disclosures(disclosures: pd.DataFrame | None) -> pd.DataFrame:
    if disclosures is None or disclosures.empty:
        return pd.DataFrame(columns=DISCLOSURE_COLUMNS)
    normalized = disclosures.copy()
    for column in DISCLOSURE_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = ""
    return normalized[DISCLOSURE_COLUMNS].fillna("").astype(str)


def normalize_disclosure_logs(logs: pd.DataFrame | None) -> pd.DataFrame:
    if logs is None or logs.empty:
        return pd.DataFrame(columns=DISCLOSURE_LOG_COLUMNS)
    normalized = logs.copy()
    for column in DISCLOSURE_LOG_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = ""
    return normalized[DISCLOSURE_LOG_COLUMNS].fillna("").astype(str)


def normalize_disclosure_watchlist(watchlist: pd.DataFrame | None) -> pd.DataFrame:
    if watchlist is None or watchlist.empty:
        return pd.DataFrame(columns=DISCLOSURE_WATCHLIST_COLUMNS)
    normalized = watchlist.copy()
    for column in DISCLOSURE_WATCHLIST_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = ""
    normalized["시장"] = normalized["시장"].map(_normalize_disclosure_market)
    normalized["티커_또는_종목코드"] = normalized["티커_또는_종목코드"].fillna("").astype(str).str.strip().str.upper()
    normalized["추적상태"] = normalized["추적상태"].replace("", "추적중")
    return normalized[DISCLOSURE_WATCHLIST_COLUMNS].fillna("").astype(str)


def _normalize_disclosure_market(value: str) -> str:
    text = str(value or "").strip().upper()
    if text in {"US", "미국"}:
        return "US"
    if text in {"KR", "한국", "KOREA"}:
        return "KR"
    return text


def _format_symbol_columns_as_text(workbook) -> None:
    target_header = "티커 또는 종목코드"
    for sheet_name in ["holdings", "prices", "calculated", "transactions", "disclosures", "disclosure_watchlist"]:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        column_index = None
        for cell in ws[1]:
            if cell.value == target_header:
                column_index = cell.column
                break
        if column_index is None:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=column_index)
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.number_format = "@"


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 35)


def is_file_locked(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        with path.open("a+b"):
            return False
    except PermissionError:
        return True
